from openpyxl import Workbook
from openpyxl.styles import PatternFill

def save_schedule_to_excel(schedule, shifts, PROVIDER_CASELOAD, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    headers = ["Date", "Day"] + list(shifts.keys()) + ["TALLY"]
    ws.append(headers)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
    for date in sorted(schedule):
        row = [date, date.strftime("%a")]
        daily_tally = 0
        for shift in shifts:
            assigned = schedule[date].get(shift, "UNASSIGNED")
            caseload = PROVIDER_CASELOAD.get(assigned, 0) if assigned != "UNASSIGNED" else 0
            row.append(f"{assigned} ({caseload})" if assigned != "UNASSIGNED" else "UNASSIGNED")
            daily_tally += caseload
        row.append(daily_tally)
        ws.append(row)
    wb.save(output_path)
