from ortools.sat.python import cp_model
from datetime import datetime, date
import logging
from openpyxl import load_workbook
import os
from datetime import datetime


class ScheduleGenerator:
    def __init__(self, employees=None, dates=None, shifts=None):
        self.employees = employees if employees else {}
        self.dates = sorted(list(dates)) if dates else []
        self.shifts = shifts if shifts else []

        # Filter out December 31st of previous year if it exists
        self.dates = [d for d in self.dates if not (isinstance(d, date) and d.month == 12 and d.day == 31)]

    def process_employee_file(self, file_path):
        """Process an employee preference Excel file in the specific format provided."""
        try:
            logging.info(f"Processing file: {file_path}")
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active

            # Get employee name from cell B4 (fallback to filename)
            emp_name = ws['B4'].value
            if not emp_name:
                emp_name = os.path.splitext(os.path.basename(file_path))[0]
            
            logging.info(f"Processing preferences for: {emp_name}")

            if emp_name not in self.employees:
                self.employees[emp_name] = {
                    'preferences': {},
                    'shift_ranges': {}
                }

            # Get shift names from row 7 (C7:K7)
            shift_names = [ws.cell(row=7, column=col).value for col in range(3, 12)]
            logging.debug(f"Shift names for {emp_name}: {shift_names}")

            # Extract shift preference ranges
            try:
                min_shifts = ws['B42'].value
                ideal_shifts = ws['B43'].value
                max_shifts = ws['B44'].value
                
                logging.info(f"{emp_name} shift ranges - Min: {min_shifts}, Ideal: {ideal_shifts}, Max: {max_shifts}")
                
                # Validate that we have at least min and max
                if min_shifts is None or max_shifts is None:
                    logging.error(f"CRITICAL: {emp_name} in file {file_path} has missing shift range values!")
                    logging.error(f"  Min shifts (B42): {min_shifts}")
                    logging.error(f"  Ideal shifts (B43): {ideal_shifts}")
                    logging.error(f"  Max shifts (B44): {max_shifts}")
                    # Set defaults to prevent errors
                    min_shifts = 0 if min_shifts is None else min_shifts
                    max_shifts = 31 if max_shifts is None else max_shifts  # Default to max days in month
                    logging.warning(f"  Using defaults for {emp_name}: Min={min_shifts}, Max={max_shifts}")
                
                self.employees[emp_name]['shift_ranges'] = {
                    'min': int(min_shifts) if min_shifts is not None else 0,
                    'ideal': int(ideal_shifts) if ideal_shifts is not None else None,
                    'max': int(max_shifts) if max_shifts is not None else 31
                }
            except Exception as e:
                logging.error(f"ERROR reading shift ranges for {emp_name} in file {file_path}: {str(e)}")
                self.employees[emp_name]['shift_ranges'] = {'min': 0, 'ideal': None, 'max': 31}

            # Get month and year from B5 (e.g., "January 2024")
            month_year_str = ws['B5'].value
            logging.debug(f"{emp_name} month/year: {month_year_str}")
            
            try:
                # Parse like "January 2024"
                my_dt = datetime.strptime(str(month_year_str).strip(), '%B %Y')
                default_month = my_dt.month
                year = my_dt.year
            except Exception as e:
                logging.error(f"Invalid month/year format in {file_path} B5 ('{month_year_str}'): {str(e)}")
                wb.close()
                return False

            # Process each day starting from row 9 to 39 (inclusive)
            preferences_count = 0
            for row_num in range(9, 40):
                day_cell = ws.cell(row=row_num, column=1).value
                if not day_cell:
                    continue

                # Parse the day (and possibly month) from column A
                try:
                    if isinstance(day_cell, datetime):
                        day = day_cell.day
                        month = day_cell.month
                    elif isinstance(day_cell, str):
                        s = day_cell.strip()
                        try:
                            dtmp = datetime.strptime(s, '%d-%b')
                            day = dtmp.day
                            month = dtmp.month
                        except ValueError:
                            if s.isdigit():
                                day = int(s)
                                month = default_month
                            else:
                                logging.warning(f"Invalid day format in {file_path} row {row_num}: {day_cell}")
                                continue
                    elif isinstance(day_cell, (int, float)):
                        day = int(day_cell)
                        month = default_month
                    else:
                        logging.warning(f"Unsupported day type in {file_path} row {row_num}: {type(day_cell)}")
                        continue
                except Exception as e:
                    logging.warning(f"Could not parse day in {file_path} row {row_num}: {str(e)}")
                    continue

                # Build the full date
                try:
                    current_date = date(year, month, day)
                except Exception as e:
                    logging.warning(f"Invalid date in {file_path} row {row_num}: y={year}, m={month}, d={day} ({str(e)})")
                    continue

                # Track date
                if current_date not in self.dates:
                    self.dates.append(current_date)

                # Initialize preference dict for this date
                if current_date not in self.employees[emp_name]['preferences']:
                    self.employees[emp_name]['preferences'][current_date] = {}

                # Read preferences for each shift in this row
                for col_num, shift in enumerate(shift_names, start=3):
                    if not shift:
                        continue
                    pref_value = ws.cell(row=row_num, column=col_num).value
                    if pref_value in {1, 2, 3}:
                        self.employees[emp_name]['preferences'][current_date][shift] = int(pref_value)
                        preferences_count += 1

            logging.info(f"Successfully processed {emp_name}: {preferences_count} preferences loaded")
            
            wb.close()
            self.dates.sort()
            return True

        except Exception as e:
            logging.error(f"FAILED to process {file_path}: {str(e)}")
            logging.exception("Full traceback:")
            return False

    SHIFT_TIMES = {
        "St. John's": "06:00",
        "Charlottetown": "08:00",
        "Halifax": "09:30",
        "Montreal": "11:30",
        "Ottawa": "14:00",
        "Toronto": "16:01",
        "Edmonton": "18:01",
        "Vancouver": "20:01",
        "Victoria": "23:00"
    }
    
    PROVIDER_CASELOAD= {
        "Adamson": 40,
        "Chan": 30,
        "Charnish": 18,
        "Da Qi": 50,
        "Duerksen": 45,
        "Duic": 25,
        "Feng": 55,
        "Gannage": 40,
        "Grenier": 55,
        "Ho": 45,
        "Hui": 45,
        "Huynh": 20,
        "Ikonnikov": 45,
        "iMak": 45,
        "Kanhai": 45,
        "Lau": 45,
        "Lee": 55,
        "Leung": 65,
        "Puri": 110,
        "Shin": 35,
        "Shum": 25,
        "Silverstein": 35,
        "Sommer": 55,
        "Shankar": 25,
        "Tan": 35,
        "Van Heer": 30,
        "Waghmare": 40,
        "Yue": 30,
    }
    
    
    def get_shift_start(self, shift_name):
        return datetime.strptime(self.SHIFT_TIMES[shift_name], "%H:%M")

    def generate_schedule(self):
        logging.info("Starting schedule generation...")
        logging.info(f"Total physicians: {len(self.employees)}")
        logging.info(f"Total dates: {len(self.dates)}")
        logging.info(f"Date range: {self.dates[0] if self.dates else 'N/A'} to {self.dates[-1] if self.dates else 'N/A'}")
        
        # Validate all employee data before starting
        for emp, emp_data in self.employees.items():
            ranges = emp_data.get('shift_ranges', {})
            min_val = ranges.get('min')
            max_val = ranges.get('max')
            
            if min_val is None or max_val is None:
                logging.error(f"CRITICAL ERROR: {emp} has invalid shift ranges!")
                logging.error(f"  Min: {min_val}, Max: {max_val}")
                logging.error(f"  This will cause the scheduler to fail!")
                raise ValueError(f"Physician {emp} has missing shift range values. Please check their preference file.")
        
        model = cp_model.CpModel()
        assignments = {}
        slack_vars = {}

        # 1) Create assignment variables for all employees for all shifts, skip Charlottetown
        for emp, emp_data in self.employees.items():
            for date_obj in self.dates:
                for shift in self.shifts:
                    if shift == "Charlottetown":
                        continue  # optional/disabled shift
                    pref = emp_data.get('preferences', {}).get(date_obj, {}).get(shift)
                    if pref in {1, 2, 3}:
                        assignments[(emp, date_obj, shift)] = model.NewBoolVar(f'{emp}_{date_obj}_{shift}')
                    else:
                        assignments[(emp, date_obj, shift)] = model.NewConstant(0)

        # 2) Force Puri’s preferred shifts if available
        if "Puri" in self.employees:
            for date_obj in self.dates:
                for shift, pref in self.employees["Puri"].get('preferences', {}).get(date_obj, {}).items():
                    if pref in {1, 2, 3} and shift != "Charlottetown":
                        key = ("Puri", date_obj, shift)
                        if key in assignments:
                            model.Add(assignments[key] == 1)

        # 3) No double booking of a shift on a day
        for date_obj in self.dates:
            for shift in self.shifts:
                if shift == "Charlottetown":
                    continue
                shift_vars = [var for (e, d, s), var in assignments.items() if d == date_obj and s == shift]
                if shift_vars:
                    model.AddAtMostOne(shift_vars)

        # 4) No physician works multiple shifts per day
        workday = {}  # Bool: employee works on this date
        for emp in self.employees:
            for date_obj in self.dates:
                daily_vars = [var for (e, d, _), var in assignments.items() if e == emp and d == date_obj]
                if daily_vars:
                    model.AddAtMostOne(daily_vars)
                    # Channel a "works today" indicator (equals sum since <=1)
                    w = model.NewBoolVar(f'work_{emp}_{date_obj}')
                    model.Add(sum(daily_vars) == w)
                    workday[(emp, date_obj)] = w
                else:
                    # If no variables exist for that emp/date, set constant 0
                    workday[(emp, date_obj)] = model.NewConstant(0)

        # 5) Puri & Lee cannot be within 8 hours of each other (same day)
        if "Puri" in self.employees and "Lee" in self.employees:
            eight_hours = 8 * 3600
            for date_obj in self.dates:
                for shift1 in self.shifts:
                    if shift1 not in self.SHIFT_TIMES or shift1 == "Charlottetown":
                        continue
                    t1 = self.get_shift_start(shift1)
                    for shift2 in self.shifts:
                        if shift2 not in self.SHIFT_TIMES or shift2 == "Charlottetown":
                            continue
                        t2 = self.get_shift_start(shift2)
                        if abs((t2 - t1).total_seconds()) <= eight_hours:
                            k1 = ("Puri", date_obj, shift1)
                            k2 = ("Lee", date_obj, shift2)
                            if k1 in assignments and k2 in assignments:
                                model.Add(assignments[k1] + assignments[k2] <= 1)

        # 6) Min/max shifts per physician (soft min via slack)
        for emp, emp_data in self.employees.items():
            emp_vars = [var for (e, _, _), var in assignments.items() if e == emp]
            if not emp_vars:
                continue
            ranges = emp_data.get('shift_ranges', {})
            min_shifts = ranges.get('min', 0) or 0
            max_shifts = ranges.get('max', len(emp_vars))

            slack = model.NewIntVar(0, max(0, min_shifts), f'slack_{emp}')
            slack_vars[emp] = slack
            model.Add(sum(emp_vars) + slack >= min_shifts)
            model.Add(sum(emp_vars) <= max_shifts)

        # 7) Max 5 consecutive workdays (SOFT): window-of-6 sum <= 5 + 6*y; penalize y in objective
        consecutive_viols = []  # collect violation indicators
        for emp in self.employees:
            n = len(self.dates)
            for i in range(0, max(0, n - 6 + 1)):
                window = [workday[(emp, self.dates[j])] for j in range(i, i + 6)]
                y = model.NewBoolVar(f'cons6_violation_{emp}_{i}')
                # If y=0 -> sum(window) <= 5 (no violation). If y=1 -> allows violation but penalized.
                model.Add(sum(window) <= 5 + 6 * y)
                consecutive_viols.append(y)

        # 8) No short turnaround (<12h) between consecutive days (SOFT)
        twelve_hours = 12
        short_turn_viols = []  # collect violation indicators for objective
        for emp in self.employees:
            for i in range(len(self.dates) - 1):
                d1 = self.dates[i]
                d2 = self.dates[i + 1]
                for s1 in self.shifts:
                    if s1 == "Charlottetown" or s1 not in self.SHIFT_TIMES:
                        continue
                    h1 = datetime.strptime(self.SHIFT_TIMES[s1], "%H:%M").hour + \
                        datetime.strptime(self.SHIFT_TIMES[s1], "%H:%M").minute / 60.0
                    for s2 in self.shifts:
                        if s2 == "Charlottetown" or s2 not in self.SHIFT_TIMES:
                            continue
                        h2 = datetime.strptime(self.SHIFT_TIMES[s2], "%H:%M").hour + \
                            datetime.strptime(self.SHIFT_TIMES[s2], "%H:%M").minute / 60.0
                        rest_hours = (24.0 - h1) + h2
                        if rest_hours < twelve_hours:
                            k1 = (emp, d1, s1)
                            k2 = (emp, d2, s2)
                            if k1 in assignments and k2 in assignments:
                                z = model.NewBoolVar(f'short_turn_{emp}_{d1}_{s1}_{d2}_{s2}')
                                # z = 1 iff both assignments are 1
                                model.Add(z <= assignments[k1])
                                model.Add(z <= assignments[k2])
                                model.Add(z >= assignments[k1] + assignments[k2] - 1)
                                short_turn_viols.append(z)

        # 9) Global total target (soft) — by default counts #assignments.
        #    Flip USE_CASELOAD_TARGET=True to target weighted caseload instead.
        USE_CASELOAD_TARGET = False
        global_target = 350
        if USE_CASELOAD_TARGET:
            weighted_terms = []
            for (emp, d, s), var in assignments.items():
                weighted_terms.append(self.PROVIDER_CASELOAD.get(emp, 0) * var)
            total_measure = sum(weighted_terms)
            global_caseload_slack = model.NewIntVar(0, max(1, global_target * 10), 'global_caseload_slack')
            model.Add(total_measure + global_caseload_slack >= global_target)
        else:
            total_assigned = sum(var for var in assignments.values())
            global_caseload_slack = model.NewIntVar(0, global_target, 'global_caseload_slack')
            model.Add(total_assigned + global_caseload_slack >= global_target)

        # 10) Vancouver fairness: evenly split among all physicians EXCEPT Puri
        vancouver_range_terms = []
        if "Vancouver" in self.shifts:
            vancouver_counts = {}
            eligible_emps = [e for e in self.employees if e != "Puri"]
            max_cap = len([d for d in self.dates])  # safe upper bound

            # count Vancouver assignments per eligible emp
            for emp in eligible_emps:
                v_vars = []
                for d in self.dates:
                    key = (emp, d, "Vancouver")
                    if key in assignments:
                        v_vars.append(assignments[key])
                if v_vars:
                    v_count = model.NewIntVar(0, len(v_vars), f'vancouver_count_{emp}')
                    model.Add(v_count == sum(v_vars))
                    vancouver_counts[emp] = v_count

            # minimize (v_max - v_min) to even the split
            if vancouver_counts:
                v_max = model.NewIntVar(0, max_cap, 'vancouver_max')
                v_min = model.NewIntVar(0, max_cap, 'vancouver_min')
                for v in vancouver_counts.values():
                    model.Add(v <= v_max)
                    model.Add(v >= v_min)
                # We'll add -(v_max - v_min) to the objective (penalty)
                vancouver_range = model.NewIntVar(0, max_cap, 'vancouver_range')
                model.Add(vancouver_range == v_max - v_min)
                vancouver_range_terms.append(vancouver_range)

        # 11) Objective: maximize preferences; penalize unmet mins, global shortfall,
        #     short turnarounds, consecutive-day violations, and Vancouver imbalance.
        objective_terms = []

        for (emp, date_obj, shift), var in assignments.items():
            if emp == "Puri":
                continue  # Puri’s chosen prefs are forced already
            pref = self.employees[emp].get('preferences', {}).get(date_obj, {}).get(shift)
            if pref in {1, 2, 3}:
                weight = {1: 3, 2: 2, 3: 1}[pref]
                objective_terms.append(weight * var)

        # penalties
        for emp, slack in slack_vars.items():
            objective_terms.append(-10 * slack)                 # unmet min shifts (strong)
        objective_terms.append(-10 * global_caseload_slack)      # below global target (strong)
        for z in short_turn_viols:
            objective_terms.append(-8 * z)                      # short turnaround violations (soft)
        for y in consecutive_viols:
            objective_terms.append(-6 * y)                      # >5 consecutive workdays (soft)
        for rng in vancouver_range_terms:
            objective_terms.append(-12 * rng)                   # Vancouver split imbalance (strong-ish)

        model.Maximize(sum(objective_terms))

        # 12) Solve
        solver = cp_model.CpSolver()
        status = solver.Solve(model)

        # 13) Build schedule with both TALLY (#assigned) and CASELOAD (sum of provider weights)
        schedule = {}
        counts = {emp: 0 for emp in self.employees}
        total_assigned_out = 0
        total_caseload_out = 0

        for date_obj in self.dates:
            schedule[date_obj] = {shift: "UNASSIGNED" for shift in self.shifts}
            tally = 0
            daily_caseload = 0
            for shift in self.shifts:
                if shift == "Charlottetown":
                    continue
                for emp in self.employees:
                    key = (emp, date_obj, shift)
                    if key in assignments and solver.Value(assignments[key]) == 1:
                        schedule[date_obj][shift] = emp
                        counts[emp] += 1
                        tally += 1
                        daily_caseload += self.PROVIDER_CASELOAD.get(emp, 0)
            schedule[date_obj]['TALLY'] = tally              # number of filled shifts that day
            schedule[date_obj]['CASELOAD'] = daily_caseload  # sum of providers' caseload weights that day
            total_assigned_out += tally
            total_caseload_out += daily_caseload

        # Optional: quick console printout (keep/remove as you prefer)
        for date_obj in sorted(schedule.keys()):
            print(f"{date_obj.strftime('%Y-%m-%d')}:")
            for shift, emp in schedule[date_obj].items():
                print(f"  {shift}: {emp}")
            print()

        stats = {
            'status': solver.StatusName(status),
            'physician_shifts': counts,
            'slack': {emp: solver.Value(slack) for emp, slack in slack_vars.items()},
            'global_caseload_slack': solver.Value(global_caseload_slack),
            'total_assigned': total_assigned_out,
            'total_caseload': total_caseload_out,
            'short_turnarounds': sum(solver.Value(z) for z in short_turn_viols),
            'consecutive_violations': sum(solver.Value(y) for y in consecutive_viols),
        }

        return schedule, stats
