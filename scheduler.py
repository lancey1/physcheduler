import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import logging
from schedule_generator import ScheduleGenerator
from ui_components import FileSelectionFrame, OutputLocationFrame, ControlFrame, SchedulePreviewFrame


class ShiftSchedulerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Physician Shift Scheduler")
        self.root.geometry("1100x1000")

        # Shift definitions
        self.shifts = {
            "St. John's": "06:00",
            "Charlottetown": "08:00",
            "Halifax": "09:30",
            "Montreal": "11:30",
            "Ottawa": "14:01",
            "Toronto": "16:01",
            "Edmonton": "18:01",
            "Vancouver": "20:01",
            "Victoria": "23:01"
        }

        self.PROVIDER_CASELOAD = {
            "Adamson": 40, "Chan": 30, "Charnish": 18, "Da Qi": 50,
            "Duerksen": 45, "Duic": 25, "Feng": 55, "Gannage": 40,
            "Grenier": 55, "Ho": 45, "Hui": 45, "Huynh": 20,
            "Ikonnikov": 45, "iMak": 45, "Kanhai": 45, "Lau": 45,
            "Lee": 55, "Leung": 65, "Puri": 110, "Shin": 35,
            "Shum": 25, "Silverstein": 35, "Sommer": 55,
            "Sethuraman": 25, "Tan": 35, "Van Heer": 30,
            "Waghmare": 40, "Yue": 30,
        }

        self.employees = {}
        self.dates = set()
        self.files = []
        self.schedule = {}
        self.pre_assignments = {}  # populated by PreAssignmentWindow or xlsx loader

        self.setup_logging()
        self.create_widgets()

    # ------------------------------------------------------------------ #
    #  Logging                                                            #
    # ------------------------------------------------------------------ #
    def setup_logging(self):
        logging.basicConfig(
            filename='scheduler.log',
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        console = logging.StreamHandler()
        console.setLevel(logging.DEBUG)
        logging.getLogger().addHandler(console)

    # ------------------------------------------------------------------ #
    #  Widget creation                                                    #
    # ------------------------------------------------------------------ #
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        self.file_frame = FileSelectionFrame(main_frame, self)
        self.file_frame.pack(fill=tk.X, pady=5)

        self.output_frame = OutputLocationFrame(main_frame, self)
        self.output_frame.pack(fill=tk.X, pady=5)

        self.control_frame = ControlFrame(main_frame, self)
        self.control_frame.pack(fill=tk.X, pady=5)

        self.results_frame = SchedulePreviewFrame(main_frame, self)
        self.results_frame.pack(fill=tk.BOTH, expand=True)

    # ------------------------------------------------------------------ #
    #  Load employee data (without generating a schedule)                 #
    # ------------------------------------------------------------------ #
    def load_employee_data_from_files(self):
        """Parse all selected preference files and populate self.employees / self.dates."""
        if not self.files:
            return False

        try:
            generator = ScheduleGenerator({}, set(), list(self.shifts.keys()))

            for file in self.files:
                if not generator.process_employee_file(file):
                    logging.warning(f"Failed to process file: {file}")

            self.employees = generator.employees
            self.dates = set(generator.dates)

            logging.info(f"Loaded {len(self.employees)} physicians, "
                         f"{len(self.dates)} dates from {len(self.files)} files")
            return True

        except Exception as e:
            logging.error(f"Failed to load employee data: {str(e)}")
            return False

    # ------------------------------------------------------------------ #
    #  Generate schedule                                                  #
    # ------------------------------------------------------------------ #
    def run_scheduler(self):
        if not self.files:
            messagebox.showerror("Error", "Please select at least one input file")
            return

        output_path = self.output_frame.output_var.get()
        if not output_path:
            messagebox.showerror("Error", "Please specify an output file location")
            return

        try:
            self.control_frame.status_var.set("Loading employee files…")
            self.root.update()

            # ── Always re-parse files so preferences are fresh ───────── #
            generator = ScheduleGenerator({}, set(), list(self.shifts.keys()))

            for file in self.files:
                if not generator.process_employee_file(file):
                    logging.warning(f"Failed to process file: {file}")

            self.employees = generator.employees
            self.dates = set(generator.dates)

            if not self.validate_data():
                return

            # ── Rebuild generator with full data ─────────────────────── #
            generator = ScheduleGenerator(self.employees, self.dates, list(self.shifts.keys()))

            # ── Inject pre-assignments ────────────────────────────────── #
            if self.pre_assignments:
                generator.set_pre_assignments(self.pre_assignments)
                total_pa = sum(len(v) for v in self.pre_assignments.values())
                logging.info(f"Injecting {total_pa} pre-assignments into solver")

            self.control_frame.status_var.set("Generating schedule (solver running)…")
            self.root.update()

            self.schedule, stats = generator.generate_schedule()

            self.results_frame.display_schedule(self.schedule)
            self.save_schedule_to_excel(self.schedule, output_path)

            status_msg = (
                f"Done – {stats['total_assigned']} shifts assigned | "
                f"Status: {stats['status']} | "
                f"Short turnarounds: {stats['short_turnarounds']}"
            )
            self.control_frame.status_var.set(status_msg)
            messagebox.showinfo("Success", f"Schedule saved to:\n{output_path}")

        except Exception as e:
            logging.exception("Scheduler failed")
            self.control_frame.status_var.set("Error occurred – see scheduler.log")
            messagebox.showerror("Error", f"Failed to generate schedule:\n{str(e)}")

    # ------------------------------------------------------------------ #
    #  Validation                                                         #
    # ------------------------------------------------------------------ #
    def validate_data(self):
        if not self.employees:
            messagebox.showerror("Error", "No physician data found in files")
            return False
        if not self.dates:
            messagebox.showerror("Error", "No valid dates found in files")
            return False
        return True

    # ------------------------------------------------------------------ #
    #  Save to Excel                                                      #
    # ------------------------------------------------------------------ #
    def save_schedule_to_excel(self, schedule, output_path):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Schedule"

            headers = ["Date", "Day"] + list(self.shifts.keys()) + ["TALLY"]
            ws.append(headers)

            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            assigned_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            unassigned_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            for cell in ws[1]:
                cell.fill = header_fill

            for date in sorted(schedule):
                row = [date, date.strftime('%a')]
                daily_tally = 0
                for shift in self.shifts:
                    assigned = schedule[date].get(shift, "UNASSIGNED")
                    if assigned != "UNASSIGNED":
                        caseload = self.PROVIDER_CASELOAD.get(assigned, 0)
                        row.append(f"{assigned} ({caseload})")
                        daily_tally += caseload
                    else:
                        row.append(assigned)
                row.append(daily_tally)
                ws.append(row)

                last_row = ws.max_row
                ws.cell(row=last_row, column=1).number_format = 'YYYY-MM-DD'

                for col_idx in range(3, len(headers) + 1):
                    cell = ws.cell(row=last_row, column=col_idx)
                    if "UNASSIGNED" in str(cell.value):
                        cell.fill = unassigned_fill
                    else:
                        cell.fill = assigned_fill

            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2

            wb.save(output_path)
            logging.info(f"Schedule saved to {output_path}")

        except Exception as e:
            logging.error(f"Failed to save Excel file: {str(e)}")
            raise

    # ------------------------------------------------------------------ #
    #  Save edited schedule                                               #
    # ------------------------------------------------------------------ #
    def save_edited_schedule(self):
        if not self.schedule:
            messagebox.showwarning("Warning", "No schedule to save. Please generate a schedule first.")
            return

        output_path = self.output_frame.output_var.get()
        if not output_path:
            messagebox.showerror("Error", "Please specify an output file location")
            return

        try:
            self.save_schedule_to_excel(self.schedule, output_path)
            self.control_frame.status_var.set("Edited schedule saved")
            messagebox.showinfo("Success", f"Edited schedule saved to:\n{output_path}")
        except Exception as e:
            logging.error(f"Failed to save edited schedule: {str(e)}")
            messagebox.showerror("Error", f"Failed to save edited schedule:\n{str(e)}")

    # ------------------------------------------------------------------ #
    #  Load existing schedule from xlsx                                   #
    # ------------------------------------------------------------------ #
    def load_schedule(self):
        file_path = filedialog.askopenfilename(
            title="Load Existing Schedule",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        if self.files and not self.employees:
            self.load_employee_data_from_files()

        try:
            from openpyxl import load_workbook

            self.control_frame.status_var.set("Loading schedule…")
            self.root.update()

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            self.schedule = {}

            headers = [cell.value for cell in ws[1] if cell.value]
            shift_columns = [
                (i, h) for i, h in enumerate(headers)
                if h not in ["Date", "Day", "TALLY"]
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                date_value = row[0]
                if isinstance(date_value, datetime):
                    date_obj = date_value.date()
                elif isinstance(date_value, str):
                    try:
                        date_obj = datetime.strptime(date_value, '%Y-%m-%d').date()
                    except Exception:
                        continue
                else:
                    continue

                self.schedule[date_obj] = {}
                for col_idx, shift_name in shift_columns:
                    cell_value = row[col_idx] if col_idx < len(row) else None
                    if cell_value and isinstance(cell_value, str):
                        physician = cell_value.split("(")[0].strip() if "(" in cell_value else cell_value.strip()
                        self.schedule[date_obj][shift_name] = physician
                    else:
                        self.schedule[date_obj][shift_name] = "UNASSIGNED"

            wb.close()
            self.output_frame.output_var.set(file_path)
            self.results_frame.display_schedule(self.schedule)
            self.control_frame.status_var.set("Schedule loaded successfully")
            messagebox.showinfo("Success", f"Schedule loaded from:\n{file_path}")
            logging.info(f"Loaded schedule from {file_path}")

        except Exception as e:
            logging.exception("Failed to load schedule")
            self.control_frame.status_var.set("Error loading schedule")
            messagebox.showerror("Error", f"Failed to load schedule:\n{str(e)}")

    # ------------------------------------------------------------------ #
    #  Load pre-assignments from an xlsx file (legacy)                    #
    # ------------------------------------------------------------------ #
    def set_pre_assignments_from_file(self):
        file_path = filedialog.askopenfilename(
            title="Load Pre-Assignments",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            self.pre_assignments = {}

            headers = [cell.value for cell in ws[1] if cell.value]
            shift_columns = [
                (i, h) for i, h in enumerate(headers)
                if h not in ["Date", "Day", "TALLY"] and h in self.shifts
            ]

            assignments_loaded = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                date_value = row[0]
                if isinstance(date_value, datetime):
                    date_obj = date_value.date()
                elif isinstance(date_value, str):
                    try:
                        date_obj = datetime.strptime(date_value, '%Y-%m-%d').date()
                    except Exception:
                        continue
                else:
                    continue

                for col_idx, shift_name in shift_columns:
                    cell_value = row[col_idx] if col_idx < len(row) else None
                    if cell_value and isinstance(cell_value, str) and cell_value != "UNASSIGNED":
                        physician = cell_value.split("(")[0].strip() if "(" in cell_value else cell_value.strip()
                        self.pre_assignments.setdefault(date_obj, {})[shift_name] = physician
                        assignments_loaded += 1

            wb.close()
            messagebox.showinfo("Success", f"Loaded {assignments_loaded} pre-assignments from:\n{file_path}")
            logging.info(f"Loaded {assignments_loaded} pre-assignments from {file_path}")

        except Exception as e:
            logging.exception("Failed to load pre-assignments")
            messagebox.showerror("Error", f"Failed to load pre-assignments:\n{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ShiftSchedulerApp(root)
    root.mainloop()